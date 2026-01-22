#!/usr/bin/env python3
"""
Inspect Excel file structure
"""
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl already installed")

def inspect_excel(excel_path):
    """Inspect Excel file structure"""
    print(f"Inspecting: {excel_path}\n")
    
    # Load workbook
    wb = openpyxl.load_workbook(excel_path)
    
    print(f"📊 Workbook: {excel_path.name}")
    print(f"   Sheets: {len(wb.sheetnames)}\n")
    
    # Inspect each sheet
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"📄 Sheet: '{sheet_name}'")
        print(f"   Dimensions: {sheet.dimensions}")
        print(f"   Max row: {sheet.max_row}")
        print(f"   Max column: {sheet.max_column}")
        
        # Show first 5 rows
        print(f"\n   First 5 rows:")
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=5, values_only=True), 1):
            # Clean up row display
            clean_row = [str(cell) if cell is not None else '' for cell in row]
            print(f"   Row {i}: {clean_row}")
        
        print()

if __name__ == "__main__":
    # Path
    project_root = Path(__file__).parent.parent
    excel_file = project_root / "data" / "AEM_Sites_Optimizer-CustomerExperience.xlsx"
    
    if not excel_file.exists():
        print(f"❌ Error: Excel file not found at {excel_file}")
        sys.exit(1)
    
    # Inspect
    inspect_excel(excel_file)
