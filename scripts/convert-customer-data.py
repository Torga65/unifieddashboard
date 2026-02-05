#!/usr/bin/env python3
"""
Convert Customer Experience Excel to JSON for EDS
Processes all weekly sheets and creates a complete dataset

Source Excel file in SharePoint:
https://adobe.sharepoint.com/sites/AEMSitesOptimizerTeam/Shared%20Documents/Unified%20Dashboard/clients/AEM_Sites_Optimizer-CustomerExperience.xlsx

Usage:
1. Download Excel from SharePoint to /data/ folder
2. Run: python3 scripts/convert-customer-data.py
3. Output: data/customers.json and data/weeks.json
"""
import json
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    pass

def parse_date_from_sheet_name(sheet_name):
    """Convert sheet name like '2026.01.15' to '2026-01-15'"""
    try:
        # Format: YYYY.MM.DD
        parts = sheet_name.split('.')
        if len(parts) == 3:
            return f"{parts[0]}-{parts[1]}-{parts[2]}"
    except:
        pass
    return None

def map_indicator_to_status(indicator):
    """Map color indicators to status values"""
    if not indicator:
        return "Unknown"
    
    indicator = str(indicator).lower().strip()
    
    mapping = {
        'green': 'Active',
        'yellow': 'At Risk',
        'red': 'Critical'
    }
    
    return mapping.get(indicator, indicator)

def calculate_health_score(engagement, blockers, feedback, health):
    """Calculate numeric health score from indicators"""
    scores = {
        'green': 100,
        'yellow': 50,
        'red': 25,
        '': 50
    }
    
    total = 0
    count = 0
    
    for indicator in [engagement, blockers, feedback, health]:
        if indicator:
            ind_str = str(indicator).lower().strip()
            total += scores.get(ind_str, 50)
            count += 1
    
    return int(total / count) if count > 0 else 50

def convert_excel_to_json(excel_path, json_path):
    """Convert all sheets to JSON"""
    print(f"📊 Converting: {excel_path.name}\n")
    
    wb = openpyxl.load_workbook(excel_path)
    all_data = []
    
    # Process each sheet (each represents a week)
    for sheet_name in wb.sheetnames:
        week_date = parse_date_from_sheet_name(sheet_name)
        
        if not week_date:
            print(f"⚠️  Skipping sheet '{sheet_name}' (can't parse date)")
            continue
        
        print(f"📄 Processing: {sheet_name} → {week_date}")
        sheet = wb[sheet_name]
        
        # Get headers from row 2
        headers = []
        for cell in sheet[2]:
            if cell.value:
                headers.append(str(cell.value).strip())
        
        # Process data rows (starting from row 3)
        week_count = 0
        for row in sheet.iter_rows(min_row=3, values_only=True):
            if not any(row):
                continue
            
            # Create row dict
            row_dict = {'week': week_date}
            
            for header, value in zip(headers, row):
                if value is not None and str(value).strip():
                    # Handle dates
                    if hasattr(value, 'strftime'):
                        row_dict[header] = value.strftime('%Y-%m-%d')
                    else:
                        row_dict[header] = str(value).strip()
            
            # Only include rows with company name
            if 'Company Name' in row_dict and row_dict['Company Name']:
                # Map engagement indicators to our format
                engagement = row_dict.get('Engagement', '')
                blockers_ind = row_dict.get('Blockers', '')
                feedback_ind = row_dict.get('Feedback', '')
                health_ind = row_dict.get('Health Score', '')
                
                # Create standardized record with ALL columns including Onboarded URLs
                customer_record = {
                    'week': week_date,
                    
                    # Core Information (Columns 1-11)
                    'companyName': row_dict.get('Company Name', ''),
                    'licenseType': row_dict.get('License Type', ''),
                    'industry': row_dict.get('Industry', ''),
                    'eseLead': row_dict.get('ESE Lead', ''),
                    'status': row_dict.get('Status', ''),
                    'delayReason': row_dict.get('Delay Reason', ''),
                    'closeDate': row_dict.get('Close Date', ''),
                    'onboardDate': row_dict.get('Onboard Date', ''),
                    'deploymentType': row_dict.get('Deployment Type', ''),
                    'headless': row_dict.get('Headless', ''),
                    'onboardedUrls': row_dict.get('Onboarded URL\'s', ''),
                    
                    # Engagement & Health (Columns 12-16)
                    'engagement': map_indicator_to_status(engagement),
                    'engagementRaw': engagement,
                    'blockersStatus': blockers_ind,
                    'blockers': 'Issues present' if blockers_ind and str(blockers_ind).lower() in ['yellow', 'red'] else 'None',
                    'feedbackStatus': feedback_ind,
                    'feedback': row_dict.get('Feedback', '') if isinstance(row_dict.get('Feedback', ''), str) and len(row_dict.get('Feedback', '')) > 10 else 'See engagement status',
                    'healthScoreRaw': health_ind,
                    'healthScore': calculate_health_score(engagement, blockers_ind, feedback_ind, health_ind),
                    'summary': row_dict.get('Summary of Engagement', ''),
                    
                    # Metrics
                    'opptyRealized': row_dict.get('Oppty Realized', ''),
                    
                    # Implementation Status (Columns 19-28)
                    'preflight': row_dict.get('Preflight', ''),
                    'autoOptimizeEnabled': row_dict.get('Auto-Opimize Enabled?', ''),
                    'autoOptimizeButtonPressed': row_dict.get('Auto-Optimize Button pressed by Customer?', ''),
                    'servicePrincipleDeployed': row_dict.get('Service Principle Deployed', ''),
                    'brandProfile': row_dict.get('Brand Profile', ''),
                    'aemyDeployed': row_dict.get('AEMY Deployed', ''),
                    'codeRepo': row_dict.get('Code Repo (git,gitlab,bitbucket,etc..)', ''),
                    'authImplementation': row_dict.get('Auth Implementation (IMS/SAML/Basic)', ''),
                    'workflowManager': row_dict.get('Workflow Manager (Jira/Asana)', ''),
                    'customerSelfServe': row_dict.get('Customer Self Serve', ''),
                    
                    # Metadata
                    'lastUpdated': week_date,
                }
                
                all_data.append(customer_record)
                week_count += 1
        
        print(f"   ✓ Processed {week_count} customers")
    
    print(f"\n💾 Total records: {len(all_data)}")
    print(f"   Writing to: {json_path}")
    
    # Create output
    output = {
        "data": all_data,
        "total": len(all_data),
        "generated": datetime.now().isoformat()
    }
    
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(output, f, indent=2, ensure_ascii=False)
    
    print(f"\n✅ Success!")
    
    # Show sample
    if all_data:
        print("\n📋 Sample record:")
        sample = all_data[0]
        for key in ['week', 'companyName', 'status', 'engagement', 'healthScore', 'summary']:
            if key in sample:
                value = sample[key]
                if len(str(value)) > 60:
                    value = str(value)[:60] + '...'
                print(f"   {key}: {value}")

if __name__ == "__main__":
    project_root = Path(__file__).parent.parent
    excel_file = project_root / "data" / "AEM_Sites_Optimizer-CustomerExperience.xlsx"
    json_file = project_root / "data" / "customers.json"
    
    if not excel_file.exists():
        print(f"❌ Error: Excel file not found")
        sys.exit(1)
    
    convert_excel_to_json(excel_file, json_file)
